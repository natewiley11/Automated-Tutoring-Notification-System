#Functions for TutoringExcelEntry program
#Creator: Nathaniel Wiley

import imaplib,re,datetime
from openpyxl import Workbook,load_workbook

#filters pertinent date and time information from body of email into datetime object
#runs in getAppointmentList function
def getRealDateandTime(t,HandM):#t is a string type
    d=int(re.compile(r'(\d+){1,2},').findall(t)[0])
    m=monthToNum(re.compile(r'(.*?) \d+').findall(t)[0])
    y=int(re.compile(r', (.*)').findall(t)[0])
    h,mins=stringToTime(re.compile(r'(.*) and').findall(HandM)[0])
    #hour is in military time
    return datetime.datetime(y,m,d,h,mins)

def monthToNum(month):#month is a string
    return{'January' : 1,
        'February' : 2,
        'March' : 3,
        'April' : 4,
        'May' : 5,
        'June' : 6,
        'July' : 7,
        'August' : 8,
        'September' : 9, 
        'October' : 10,
        'November' : 11,
        'December' : 12
    }[month]

#converts string to int values for time
def stringToTime(t):
    h,m=t[0:-2].split(':')
    h=int(h);m=int(m)
    if(t[-2:]=='pm'):
        h+=12
    return h,m

#enters appointment in sorted order into excel
def excelDate(TD,subject,ws):
    if TD!='fail':
        LessTD=TD-datetime.timedelta(minutes=1)
        MoreTD=TD+datetime.timedelta(minutes=1)
    else:
        return
    print('through dates')
    i=2
    if ws.cell(row=i,column=2).value==None:
        ws.cell(row=i,column=2).value=TD
        ws.cell(row=i,column=3).value=subject
        return
    while ws.cell(row=i,column=2).value!=None:
        if LessTD<=ws.cell(row=i,column=2).value<=MoreTD:
            print('same date')
            #date already entered
            break
        elif ws.cell(row=i,column=2).value > TD:
            print('entry made')   #filters most recent on top
            ws.insert_rows(i)
            ws.cell(row=i,column=2).value=TD
            ws.cell(row=i,column=3).value=subject
            break
        elif ws.cell(row=i+1,column=2).value==None:
            print('Entered last')
            ws.insert_rows(i+1)
            ws.cell(row=i+1,column=2).value=TD
            ws.cell(row=i+1,column=3).value=subject
            break
        else:
            print('iter')
            i+=1
#gathers appointment date, time and subject from a specific email
#runs one time for each email
def getAppointmentList(id_list,h,mail):
    current_email_id=id_list[h]
    result, data = mail.fetch(current_email_id, "(UID BODY[TEXT])")
    #double check to make sure its a tutoring appointment
    check, f= mail.fetch(current_email_id, "(BODY[HEADER.FIELDS (FROM)])")
    #subject line: to check made or cancelled
    c, subject= mail.fetch(current_email_id, "(BODY[HEADER.FIELDS (SUBJECT)])")
    if (result and check)=='OK':
        email=str(data[0][1])
    else:
        return ['fail','fail']#fail-safe for failed email grab
    tempInfo=re.compile(r'an appointment on (.*?) as your topic').findall(email)
    s=re.compile(r'(\\r|\\n|\\)')
    out=s.sub('',tempInfo[0])
    tempDate=re.compile(r', (.*)between').findall(out)[0]
    tempTime=re.compile(r'between(.*) and').findall(out)[0]
    tempSubject=re.compile(r'chosen (.*)').findall(out)[0]
    potentialAppointment=getRealDateandTime(tempDate,tempTime)
    if datetime.datetime.today()<=potentialAppointment:#appointment is in future
        return [potentialAppointment,tempSubject]
    else:#appoiintment in past
        print('fail',potentialAppointment)
        return ['fail','fail']
